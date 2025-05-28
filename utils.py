# utils.py

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# === Global Styles ===
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')


# === Date Formatting ===
def get_formatted_dates(date_str='20250430'):
    current_date = pd.to_datetime(date_str, format='%Y%m%d')
    last_date = current_date - pd.offsets.MonthEnd(1)

    formatted_current = f"{current_date.month}/{current_date.day}/{str(current_date.year)[-2:]}"
    formatted_last = f"{last_date.month}/{last_date.day}/{str(last_date.year)[-2:]}"

    return formatted_current, formatted_last


# === Excel Utilities ===

def get_column_index(ws, column_name):
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == column_name:
            return col[0].column
    raise KeyError(f"'{column_name}' column not found.")


def format_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = ALIGN_CENTER

def adjust_column_widths(ws, max_width_limit=60, min_width=10):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0

        for cell in col:
            if cell.value is not None:
                value = str(cell.value)
                length = len(value)
                if length > max_length:
                    max_length = length

        adjusted_width = min(max(max_length + 2, min_width), max_width_limit)
        ws.column_dimensions[col_letter].width = adjusted_width

def format_worksheet(ws):
    for col in ws.iter_cols():
        header = col[0].value
        if header and 'IRR' in header:
            for cell in col[1:]:
                cell.number_format = '0.00%'
        elif header in ['Liq Cap', 'Market Value', f'{get_current_mv_col()}']:
            for cell in col[1:]:
                cell.number_format = '#,##0_);(#,##0)'
        elif header and 'Duration' in header:
            for cell in col[1:]:
                cell.number_format = '0.00'

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = ALIGN_CENTER

    for cell in ws[1]:
        format_header_cell(cell)

    adjust_column_widths(ws)



def format_all_sheets(*sheets):
    for sheet in sheets:
        format_worksheet(sheet)


# Optional helper if needed
def get_current_mv_col():
    # Just a placeholder for f'{formatted_date} MV' if not passed in explicitly
    return 'MV'