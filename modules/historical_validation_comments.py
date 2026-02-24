"""
Copy AAT Comments from Previous Version

This module automatically copies AAT Comments from the previous version file
to the current version for Highlight IRR Diffs and Highlight Duration Diffs tabs.

Workflow:
1. Find the latest version file for the given date in the source folder
2. Find the previous version (v(n-1) or last month's last version)
3. Extract AAT Comments from previous version
4. Update AAT Comments in the latest version file
"""

import os
import re
import sys
import shutil
from pathlib import Path
from typing import Optional, Tuple, Dict, List
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# Add parent directory to path to import config
parent_dir = Path(__file__).parent.parent
sys.path.insert(0, str(parent_dir))

import config

# ===== Configuration =====
# Import source folder from config
SOURCE_FOLDER = config.VERSIONED_FILES_FOLDER

# Sheets to update
TARGET_SHEETS = ['Highlight IRR Diffs', 'Highlight Duration Diffs']


def parse_filename(filename: str) -> Optional[Tuple[str, int]]:
    """
    Parse filename to extract date and version number.

    Args:
        filename: Filename like 'AAT vs ECF 20251130.v2.xlsx'

    Returns:
        Tuple of (date_str, version_number) or None if parsing fails
    """
    # Pattern: date (8 digits) followed by .v and version number
    pattern = r'(\d{8})\.v(\d+)'
    match = re.search(pattern, filename)

    if match:
        date_str = match.group(1)
        version = int(match.group(2))
        return (date_str, version)

    return None


def find_latest_version(date_str: str, source_folder: str) -> Optional[Tuple[str, int]]:
    """
    Find the latest version file for the given date.

    Args:
        date_str: Date string in format 'YYYYMMDD'
        source_folder: Folder containing version files

    Returns:
        Tuple of (filepath, version_number) or None if not found
    """
    try:
        all_files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
    except Exception as e:
        print(f"  [Error] Cannot read folder {source_folder}: {e}")
        return None

    # Find all files matching the date
    matching_files = []
    for filename in all_files:
        parsed = parse_filename(filename)
        if parsed and parsed[0] == date_str:
            matching_files.append((filename, parsed[1]))

    if not matching_files:
        print(f"  [Error] No version files found for date {date_str}")
        return None

    # Sort by version number (descending) and get the latest
    matching_files.sort(key=lambda x: x[1], reverse=True)
    latest_file = matching_files[0]

    filepath = os.path.join(source_folder, latest_file[0])
    return (filepath, latest_file[1])


def find_previous_version(date_str: str, current_version: int, source_folder: str) -> Optional[str]:
    """
    Find the previous version file.

    Logic:
    - If current is v2 or higher -> find v(n-1) in same month
    - If current is v1 -> find last version of previous month

    Args:
        date_str: Current date string
        current_version: Current version number
        source_folder: Folder containing version files

    Returns:
        Path to previous version file, or None if not found
    """
    print(f"  - Current: {date_str}.v{current_version}")

    try:
        all_files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
    except Exception as e:
        print(f"  [Error] Cannot read folder {source_folder}: {e}")
        return None

    if current_version > 1:
        # Find v(n-1) in same month
        target_version = current_version - 1
        target_pattern = f"{date_str}.v{target_version}"

        for filename in all_files:
            if target_pattern in filename:
                prev_file = os.path.join(source_folder, filename)
                print(f"  - Previous: {date_str}.v{target_version}")
                return prev_file

        print(f"  [Error] Cannot find {date_str}.v{target_version}")
        return None

    else:
        # Current is v1, find last version of previous month
        current_dt = datetime.strptime(date_str, '%Y%m%d')
        prev_month_dt = current_dt - relativedelta(months=1)
        prev_month_str = prev_month_dt.strftime('%Y%m%d')

        # Find all files from previous month
        prev_month_files = []
        for filename in all_files:
            parsed = parse_filename(filename)
            if parsed and parsed[0] == prev_month_str:
                prev_month_files.append((filename, parsed[1]))

        if not prev_month_files:
            print(f"  [Error] No files found for previous month {prev_month_str}")
            return None

        # Get the file with highest version
        prev_month_files.sort(key=lambda x: x[1], reverse=True)
        prev_file_name = prev_month_files[0][0]
        prev_file = os.path.join(source_folder, prev_file_name)
        print(f"  - Previous: {prev_month_str}.v{prev_month_files[0][1]} (last month)")
        return prev_file


def extract_comments_mapping(file_path: str, sheet_names: List[str]) -> Dict[str, str]:
    """
    Extract Deal Name -> AAT Comments mapping from specified sheets.

    Args:
        file_path: Path to Excel file
        sheet_names: List of sheet names to extract from

    Returns:
        Dictionary mapping Deal Name to AAT Comments
    """
    comments_map = {}
    extracted_count = 0

    try:
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                print(f"    [Skip] Sheet '{sheet_name}' not found")
                continue

            ws = wb[sheet_name]

            # Find column indices
            deal_name_col = None
            comments_col = None

            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == 'Deal Name':
                    deal_name_col = col_idx
                elif header == 'AAT Comments':
                    comments_col = col_idx

            if not deal_name_col:
                print(f"    [Skip] 'Deal Name' column not found in '{sheet_name}'")
                continue

            if not comments_col:
                print(f"    [Skip] 'AAT Comments' column not found in '{sheet_name}'")
                continue

            # Extract mapping
            sheet_count = 0
            for row_idx in range(2, ws.max_row + 1):
                deal_name = ws.cell(row=row_idx, column=deal_name_col).value
                comment = ws.cell(row=row_idx, column=comments_col).value

                if deal_name and comment:
                    # Use deal name as key
                    key = str(deal_name).strip()
                    comments_map[key] = comment
                    sheet_count += 1

            if sheet_count > 0:
                print(f"    - '{sheet_name}': {sheet_count} comments")
                extracted_count += sheet_count

        wb.close()

    except Exception as e:
        print(f"    [Error] Failed to read {os.path.basename(file_path)}: {e}")
        return {}

    return comments_map


def update_comments(target_file: str, comments_map: Dict[str, str], sheet_names: List[str]) -> int:
    """
    Update AAT Comments in target file based on mapping.

    Args:
        target_file: Path to target Excel file
        comments_map: Dictionary mapping Deal Name to AAT Comments
        sheet_names: List of sheet names to update

    Returns:
        Number of cells updated
    """
    updated_count = 0

    try:
        # Load workbook and preserve as much formatting as possible
        wb = load_workbook(target_file, keep_vba=True, keep_links=True)

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                print(f"    [Skip] Sheet '{sheet_name}' not found in target file")
                continue

            ws = wb[sheet_name]

            # Find column indices
            deal_name_col = None
            comments_col = None

            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == 'Deal Name':
                    deal_name_col = col_idx
                elif header == 'AAT Comments':
                    comments_col = col_idx

            if not deal_name_col or not comments_col:
                print(f"    [Skip] Required columns not found in '{sheet_name}'")
                continue

            # Update comments
            sheet_updated = 0
            for row_idx in range(2, ws.max_row + 1):
                deal_name = ws.cell(row=row_idx, column=deal_name_col).value

                if deal_name:
                    key = str(deal_name).strip()

                    # Check if we have a comment for this deal
                    if key in comments_map:
                        ws.cell(row=row_idx, column=comments_col).value = comments_map[key]
                        sheet_updated += 1

            if sheet_updated > 0:
                print(f"    - '{sheet_name}': {sheet_updated} cells updated")
                updated_count += sheet_updated

        # Save the workbook
        wb.save(target_file)
        wb.close()

    except Exception as e:
        print(f"    [Error] Failed to update {os.path.basename(target_file)}: {e}")
        raise

    return updated_count


def save_to_summary_report(source_file: str, date_str: str, summary_folder: str) -> None:
    """
    Save a versioned copy of the updated file to the AAT_ECF_SUMMARY_REPORT folder.

    Checks existing files in the summary folder for the same date:
    - If no files exist for this date -> saves as v1
    - If last version is vN -> saves as v(N+1)

    Args:
        source_file: Path to the updated source file
        date_str: Date string in format 'YYYYMMDD'
        summary_folder: Target folder (config.AAT_ECF_SUMMARY_REPORT)
    """
    try:
        os.makedirs(summary_folder, exist_ok=True)

        # Scan existing files for this date to determine next version
        try:
            all_files = [f for f in os.listdir(summary_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
        except Exception as e:
            print(f"  [Warning] Cannot read summary folder {summary_folder}: {e}")
            return

        pattern = rf'{date_str}\.v(\d+)'
        max_version = 0
        for filename in all_files:
            match = re.search(pattern, filename)
            if match:
                version = int(match.group(1))
                max_version = max(max_version, version)

        next_version = max_version + 1

        # Copy with versioned filename
        versioned_filename = f'AAT vs ECF {date_str}.v{next_version}.xlsx'
        versioned_path = os.path.join(summary_folder, versioned_filename)
        shutil.copy2(source_file, versioned_path)

        if max_version == 0:
            print(f"  [OK] New month - saved as: {versioned_filename}")
        else:
            print(f"  [OK] Last version was v{max_version} - saved as: {versioned_filename}")
        print(f"  [OK] Path: {versioned_path}")

    except Exception as e:
        print(f"  [Warning] Failed to save to summary report: {e}")


def run_copy_comments(date_str: str) -> None:
    """
    Main function to copy comments from previous version.

    Args:
        date_str: Date string in format 'YYYYMMDD'
    """
    print("\n" + "=" * 80)
    print("Module: Copy AAT Comments from Previous Version")
    print("=" * 80)
    print(f"\nProcessing date: {date_str}")
    print(f"Source folder: {SOURCE_FOLDER}")

    # Step 1: Find latest version file for the date
    print("\n[1/5] Finding latest version file...")
    latest = find_latest_version(date_str, SOURCE_FOLDER)

    if not latest:
        return

    target_file, current_version = latest
    print(f"  - Found: {os.path.basename(target_file)}")

    # Step 2: Find previous version
    print("\n[2/5] Finding previous version...")
    prev_file = find_previous_version(date_str, current_version, SOURCE_FOLDER)

    if not prev_file:
        return

    if not os.path.exists(prev_file):
        print(f"  [Error] Previous file does not exist: {prev_file}")
        return

    print(f"  - Found: {os.path.basename(prev_file)}")

    # Step 3: Extract comments from previous version
    print("\n[3/5] Extracting comments from previous version...")
    comments_map = extract_comments_mapping(prev_file, TARGET_SHEETS)

    if not comments_map:
        print("  [Warning] No comments found in previous version")
        return

    print(f"  - Total unique comments: {len(comments_map)}")

    # Step 4: Update target file
    print("\n[4/5] Updating target file...")
    updated_count = update_comments(target_file, comments_map, TARGET_SHEETS)

    if updated_count > 0:
        print(f"  [OK] {updated_count} cells updated in: {os.path.basename(target_file)}")
    else:
        print("  [Warning] No cells were updated (all Deal Names may be new)")

    # Step 5: Save to AAT_ECF_SUMMARY_REPORT with version numbering
    print("\n[5/5] Saving to AAT ECF Summary Report...")
    print(f"  - Target folder: {config.AAT_ECF_SUMMARY_REPORT}")
    save_to_summary_report(target_file, date_str, config.AAT_ECF_SUMMARY_REPORT)

    print(f"\n{'=' * 80}")
    if updated_count > 0:
        print(f"[SUCCESS] Comments copied successfully!")
        print(f"  - File: {os.path.basename(target_file)}")
        print(f"  - Total cells updated: {updated_count}")
    else:
        print("[WARNING] No cells were updated (all Deal Names may be new)")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    # For standalone testing
    DEFAULT_DATE = '20251130'
    run_copy_comments(DEFAULT_DATE)
