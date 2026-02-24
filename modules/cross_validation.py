"""
AAT vs ECF Cross-Validation Report Generator

This module processes and compares AAT and ECF data to generate comprehensive Excel reports with highlighting and categorization.
"""

from typing import List, Tuple
import pandas as pd
import os
import sys
import re
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

# Add parent directory to path to import utils
parent_dir = Path(__file__).parent.parent
sys.path.insert(0, str(parent_dir))

from utils import get_column_index, format_header_cell, format_all_sheets
import config

# ===== Import Configuration from config.py =====
SIGNIFICANT_MV_THRESHOLD = config.SIGNIFICANT_MV_THRESHOLD
IRR_DIFF_THRESHOLD = config.IRR_DIFF_THRESHOLD
DURATION_DIFF_THRESHOLD = config.DURATION_DIFF_THRESHOLD
BASE_PATH = config.BASE_PATH
AAT_OUTPUT_BASE_PATH = config.AAT_OUTPUT_BASE_PATH

# ===== Global Cell Styles =====
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

# Highlight colors
HIGHLIGHT_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HIGHLIGHT_ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
HIGHLIGHT_GREEN = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
HIGHLIGHT_GRAY = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


def initialize_dates(date_str: str) -> Tuple[str, str]:
    """
    Initialize current and previous month dates from date string.

    Args:
        date_str: Date string in format 'YYYYMMDD'

    Returns:
        Tuple of (formatted_current_date, formatted_last_date)
    """
    current_date = pd.to_datetime(date_str, format='%Y%m%d')
    last_date = current_date - pd.offsets.MonthEnd(1)

    formatted_current = f"{current_date.month}/{current_date.day}/{str(current_date.year)[-2:]}"
    formatted_last = f"{last_date.month}/{last_date.day}/{str(last_date.year)[-2:]}"

    print(f"Processing dates: Current={formatted_current}, Previous={formatted_last}")
    return formatted_current, formatted_last


def get_file_paths(date_str: str) -> dict:
    """
    Generate all required file paths based on date string.

    Args:
        date_str: Date string in format 'YYYYMMDD'

    Returns:
        Dictionary containing all file paths
    """
    return {
        'status_final': f'{BASE_PATH}/{date_str}/Status_Final_{date_str}.xlsx',
        'aat_data': f'{AAT_OUTPUT_BASE_PATH}/{date_str}/AATOutput.{date_str}.xlsx',
        'aat_pm_owner': f'{BASE_PATH}/AAT PM Owner.xlsx',
        'output_folder': f'{BASE_PATH}/{date_str}',
        'output_filename': f'AAT vs ECF {date_str}.xlsx'
    }


def load_data(aat_path: str, status_path: str) -> pd.DataFrame:
    """
    Load and merge AAT and status data.

    Args:
        aat_path: Path to AAT data Excel file
        status_path: Path to status final Excel file

    Returns:
        Merged DataFrame

    Raises:
        FileNotFoundError: If input files don't exist
    """
    if not os.path.exists(aat_path):
        raise FileNotFoundError(f"AAT data file not found: {aat_path}")
    if not os.path.exists(status_path):
        raise FileNotFoundError(f"Status file not found: {status_path}")

    df_aat = pd.read_excel(aat_path)
    df_status = pd.read_excel(status_path)

    # Filter Status_Final to only keep Deal-level rows (where Instrument is empty)
    # This removes individual instrument rows and keeps only SUBTOTAL rows with aggregated MV
    print(f"  - Status file before filtering: {len(df_status)} rows")
    df_status = df_status[df_status['Instrument'].isna()]
    print(f"  - Status file after filtering (Deal-level only): {len(df_status)} rows")

    return pd.merge(df_aat, df_status, on='Deal Name', how='left')


def process_data(df: pd.DataFrame, pm_owner_path: str,
                 current_date: str, last_date: str) -> pd.DataFrame:
    """
    Process and transform the merged data.

    Args:
        df: Merged DataFrame
        pm_owner_path: Path to PM owner mapping file
        current_date: Formatted current date string
        last_date: Formatted previous date string

    Returns:
        Processed DataFrame with calculated columns
    """
    # Remove unnecessary columns
    df.drop(columns=['Instrument', 'Abs IRR Change'], inplace=True)
    df.drop_duplicates(subset='Deal Name', keep='first', inplace=True)

    # Calculate IRR differences
    irr_col = f'{current_date} IRR'
    aat_irr_col = f'{current_date} AAT IRR'
    df.insert(
        df.columns.get_loc(irr_col) + 1,
        'AAT&ECF IRR Diffs',
        df[irr_col] - df[aat_irr_col]
    )

    # Calculate duration differences
    df.insert(
        df.columns.get_loc('Duration DCF Base¹') + 1,
        'Duration Diffs',
        df['Duration DCF Base¹'] - df['Duration AAT Base']
    )

    # Map PM owners
    pm_map = pd.read_excel(pm_owner_path).set_index('Sr. Portfolio Manager')['AAT PM Owner']
    df.insert(
        df.columns.get_loc('Sr. Portfolio Manager') + 1,
        'AAT PM Owner',
        df['Sr. Portfolio Manager'].map(pm_map)
    )

    # Reposition Liq Cap and Market Value columns
    df_liq_cap = df.pop('Liq Cap')
    df_market_value = df.pop(f'{current_date} MV')
    df.insert(df.columns.get_loc('Duration Diffs') + 1, 'Liq Cap', df_liq_cap)
    df.insert(df.columns.get_loc('Liq Cap') + 1, f'{current_date} MV', df_market_value)

    # Rename columns for clarity
    df.rename(columns={
        f'{current_date} IRR': f'{current_date} ECF IRR',
        f'{last_date} IRR': f'{last_date} ECF IRR',
        'IRR Change': 'MoM ECF IRR Movements',
        'Duration AAT Base': 'Duration AAT',
        'Duration DCF Base¹': 'Duration ECF',
        'Comments': 'AAT Comments'
    }, inplace=True)

    return df


def reorder_columns(df: pd.DataFrame, current_date: str, last_date: str) -> pd.DataFrame:
    """
    Reorder DataFrame columns in desired sequence.

    Args:
        df: DataFrame to reorder
        current_date: Formatted current date string
        last_date: Formatted previous date string

    Returns:
        DataFrame with reordered columns
    """
    columns_order = [
        'Deal Name',
        'Sr. Portfolio Manager',
        'AAT PM Owner',
        f'{current_date} AAT IRR',
        f'{current_date} ECF IRR',
        'AAT&ECF IRR Diffs',
        f'{last_date} ECF IRR',
        'MoM ECF IRR Movements',
        'Duration AAT',
        'Duration ECF',
        'Duration Diffs',
        'Liq Cap',
        f'{current_date} MV',
        'MV %',
        'AAT Comments',
    ]
    return df[columns_order]


def calculate_metrics(df: pd.DataFrame, current_date: str) -> pd.DataFrame:
    """
    Calculate market value percentage and cumulative metrics.

    Args:
        df: DataFrame to process
        current_date: Formatted current date string

    Returns:
        DataFrame with calculated metrics
    """
    mv_col = f'{current_date} MV'
    total_mv = df[mv_col].sum()

    # Calculate MV percentage
    df['MV %'] = df[mv_col] / total_mv * 100
    df['MV %'] = df['MV %'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")

    # Sort by market value
    df.sort_values(by=mv_col, ascending=False, inplace=True)

    # Calculate cumulative MV percentage
    df['Cumulative MV %'] = df[mv_col].cumsum() / total_mv * 100
    df['Cumulative MV %'] = df['Cumulative MV %'].apply(
        lambda x: f"{x:.2f}%" if pd.notnull(x) else ""
    )

    return df


def save_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Save DataFrame to Excel file.

    Args:
        df: DataFrame to save
        output_path: Path for output Excel file
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Summary')


def highlight_and_collect(ws: Worksheet, column_name: str, threshold: float,
                          fill: PatternFill, mv_col_name: str,
                          mv_threshold: float = SIGNIFICANT_MV_THRESHOLD) -> List[List]:
    """
    Highlight cells exceeding threshold and collect significant rows.

    Args:
        ws: Worksheet to process
        column_name: Name of column to check
        threshold: Threshold value for highlighting
        fill: PatternFill to apply
        mv_col_name: Market value column name
        mv_threshold: Market value threshold for significance

    Returns:
        List of significant rows

    Raises:
        KeyError: If market value column not found
    """
    significant_rows = []

    # Find market value column index
    mv_col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == mv_col_name:
            mv_col_idx = col[0].column
            break

    if mv_col_idx is None:
        raise KeyError(f"'{mv_col_name}' column not found in worksheet")

    # Process target column
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == column_name:
            for cell in col[1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    if abs(cell.value) > threshold:
                        market_value = ws.cell(row=cell.row, column=mv_col_idx).value
                        if market_value is not None and market_value >= mv_threshold:
                            cell.fill = fill
                            row_values = [c.value for c in ws[cell.row]]
                            significant_rows.append(row_values)
            break

    return significant_rows


def identify_significant_changes(ws: Worksheet, current_date: str) -> Tuple[List, List, List]:
    """
    Identify and highlight significant changes and differences.

    Args:
        ws: Worksheet to process
        current_date: Formatted current date string

    Returns:
        Tuple of (significant_changes, significant_diffs, duration_diffs)
    """
    mv_col_name = f'{current_date} MV'

    significant_changes = highlight_and_collect(
        ws, 'MoM ECF IRR Movements', IRR_DIFF_THRESHOLD, HIGHLIGHT_YELLOW, mv_col_name
    )
    significant_diffs = highlight_and_collect(
        ws, 'AAT&ECF IRR Diffs', IRR_DIFF_THRESHOLD, HIGHLIGHT_ORANGE, mv_col_name
    )
    duration_diffs = highlight_and_collect(
        ws, 'Duration Diffs', DURATION_DIFF_THRESHOLD, HIGHLIGHT_GREEN, mv_col_name
    )

    return significant_changes, significant_diffs, duration_diffs


def format_worksheet(ws: Worksheet, current_date: str) -> None:
    """
    Apply formatting to worksheet cells and columns.

    Args:
        ws: Worksheet to format
        current_date: Formatted current date string
    """
    # Apply number formatting
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if header and 'IRR' in str(header):
            for cell in col[1:]:
                cell.number_format = '0.00%'
        elif header in [f'{current_date} MV', 'Liq Cap']:
            for cell in col[1:]:
                cell.number_format = '#,##0_);(#,##0)'
        elif header and 'Duration' in str(header):
            for cell in col[1:]:
                cell.number_format = '0.00'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Center align data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = ALIGN_CENTER

    # Format headers
    for cell in ws[1]:
        format_header_cell(cell)


def remove_columns_from_sheet(ws: Worksheet, columns_to_remove: List[str]) -> None:
    """
    Remove specified columns from a worksheet.

    Args:
        ws: Worksheet to process
        columns_to_remove: List of column names to remove
    """
    # Iterate in reverse order to avoid index shifting issues
    for col_idx in range(ws.max_column, 0, -1):
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value in columns_to_remove:
            ws.delete_cols(col_idx)


def create_highlighted_sheets(wb: Workbook, significant_changes: List,
                              significant_diffs: List, duration_diffs: List,
                              current_date: str, last_date: str) -> None:
    """
    Create separate sheets for different types of significant items.

    Args:
        wb: Workbook to add sheets to
        significant_changes: Rows with significant IRR changes
        significant_diffs: Rows with significant AAT/ECF differences
        duration_diffs: Rows with significant duration differences
        current_date: Formatted current date string
        last_date: Formatted previous date string
    """
    header = [cell.value for cell in wb['Summary'][1]]

    def create_sheet(name: str, rows: List) -> Worksheet:
        """Helper function to create and populate a sheet."""
        ws = wb.create_sheet(title=name)
        ws.append(header)
        for row in rows:
            ws.append(row)
        return ws

    # Create sheets
    ws_changes = create_sheet('Significant ECF IRR Movers', significant_changes)
    ws_diffs = create_sheet('Highlight IRR Diffs', significant_diffs)
    ws_durations = create_sheet('Highlight Duration Diffs', duration_diffs)

    # Remove columns from Highlight IRR Diffs sheet
    irr_diffs_columns_to_remove = [
        f'{last_date} ECF IRR',
        'MoM ECF IRR Movements',
        'Duration AAT',
        'Duration ECF',
        'Duration Diffs'
    ]
    remove_columns_from_sheet(ws_diffs, irr_diffs_columns_to_remove)

    # Remove all IRR-related columns from Highlight Duration Diffs sheet
    # First, identify all column headers containing 'IRR'
    duration_columns_to_remove = []
    for col in ws_durations.iter_cols(1, ws_durations.max_column):
        header = col[0].value
        if header and 'IRR' in str(header):
            duration_columns_to_remove.append(header)
    remove_columns_from_sheet(ws_durations, duration_columns_to_remove)

    format_all_sheets(ws_changes, ws_diffs, ws_durations)


def add_category_column(wb: Workbook, current_date: str) -> None:
    """
    Add categorization column to summary sheet based on both IRR and Duration differences.

    Args:
        wb: Workbook containing summary sheet
        current_date: Formatted current date string
    """
    ws = wb['Summary']

    # Insert category column after MV %
    mv_pct_col_idx = get_column_index(ws, 'MV %')
    category_col_idx = mv_pct_col_idx + 1
    ws.insert_cols(category_col_idx)

    # Format header
    cell = ws.cell(row=1, column=category_col_idx, value='Category')
    format_header_cell(cell)

    # Get column indices
    irr_diff_col_idx = get_column_index(ws, 'AAT&ECF IRR Diffs')
    duration_diff_col_idx = get_column_index(ws, 'Duration Diffs')
    mv_col_idx = get_column_index(ws, f'{current_date} MV')

    # Categorize each row based on both IRR and Duration differences
    for row in range(2, ws.max_row + 1):
        irr_diff = ws.cell(row=row, column=irr_diff_col_idx).value
        duration_diff = ws.cell(row=row, column=duration_diff_col_idx).value
        mv_value = ws.cell(row=row, column=mv_col_idx).value

        # Check if either IRR diff or Duration diff exceeds threshold
        has_irr_discrepancy = irr_diff is not None and abs(irr_diff) > IRR_DIFF_THRESHOLD
        has_duration_discrepancy = duration_diff is not None and abs(duration_diff) > DURATION_DIFF_THRESHOLD

        if irr_diff is not None or duration_diff is not None:
            if mv_value is not None and mv_value > SIGNIFICANT_MV_THRESHOLD:
                # Significant MV: categorize based on whether there's any discrepancy
                category = 'Significant Discrepancy' if (has_irr_discrepancy or has_duration_discrepancy) else 'Alignment'
            else:
                # Small MV: note discrepancy but mark as ignore
                category = 'Significant discrepancy but ignore' if (has_irr_discrepancy or has_duration_discrepancy) else 'Alignment'
            ws.cell(row=row, column=category_col_idx, value=category)


def drop_cumulative_mv_column(wb: Workbook) -> None:
    """
    Remove cumulative MV% column from all sheets except Summary.

    Args:
        wb: Workbook to process
    """
    for ws in wb.worksheets:
        if ws.title != 'Summary':
            for col in ws.iter_cols(1, ws.max_column):
                if col[0].value == 'Cumulative MV %':
                    ws.delete_cols(col[0].column)
                    break


def create_missing_aat_sheet(wb: Workbook, current_date: str) -> None:
    """
    Create a sheet listing all deals that are missing AAT IRR or Duration AAT values.

    Args:
        wb: Workbook to add the sheet to
        current_date: Formatted current date string
    """
    ws_summary = wb['Summary']
    aat_irr_col_name = f'{current_date} AAT IRR'
    duration_aat_col_name = 'Duration AAT'

    # Columns to display in the new sheet (plus a computed 'Missing Fields' column)
    display_cols = [
        'Deal Name',
        'Sr. Portfolio Manager',
        'AAT PM Owner',
        aat_irr_col_name,
        duration_aat_col_name,
        'Liq Cap',
        f'{current_date} MV',
    ]

    # Build index map from Summary header (0-based)
    header = [cell.value for cell in ws_summary[1]]

    aat_irr_idx = header.index(aat_irr_col_name) if aat_irr_col_name in header else None
    duration_aat_idx = header.index(duration_aat_col_name) if duration_aat_col_name in header else None

    if aat_irr_idx is None and duration_aat_idx is None:
        print(f"  [Skip] Neither '{aat_irr_col_name}' nor '{duration_aat_col_name}' found in Summary")
        return

    col_idx_map = {col: header.index(col) for col in display_cols if col in header}
    liq_cap_idx = header.index('Liq Cap') if 'Liq Cap' in header else None

    # Collect rows where AAT IRR or Duration AAT is missing
    missing_rows = []
    liq_cap_values = []  # parallel list used for Deal Name highlighting
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, values_only=True):
        aat_irr_missing = aat_irr_idx is not None and row[aat_irr_idx] is None
        duration_aat_missing = duration_aat_idx is not None and row[duration_aat_idx] is None

        if aat_irr_missing or duration_aat_missing:
            missing_fields = []
            if aat_irr_missing:
                missing_fields.append('AAT IRR')
            if duration_aat_missing:
                missing_fields.append('Duration AAT')

            selected = [row[col_idx_map[col]] if col in col_idx_map else None for col in display_cols]
            selected.append(', '.join(missing_fields))
            missing_rows.append(selected)
            liq_cap_values.append(row[liq_cap_idx] if liq_cap_idx is not None else None)

    # Sort by Liq Cap descending (None treated as 0)
    paired = sorted(zip(missing_rows, liq_cap_values), key=lambda x: x[1] or 0, reverse=True)
    missing_rows, liq_cap_values = (list(col) for col in zip(*paired)) if paired else ([], [])

    # Create sheet
    ws_missing = wb.create_sheet(title='Missing AAT Data')
    ws_missing.append(display_cols + ['Missing Fields'])
    for row_data in missing_rows:
        ws_missing.append(row_data)

    # Apply standard formatting
    format_all_sheets(ws_missing)

    all_cols = display_cols + ['Missing Fields']
    deal_name_display_col = all_cols.index('Deal Name') + 1          # 1-based
    aat_irr_display_col = all_cols.index(aat_irr_col_name) + 1       # 1-based
    duration_aat_display_col = all_cols.index(duration_aat_col_name) + 1  # 1-based

    # Highlight missing cells in red; highlight Deal Name in gray for large deals (Liq Cap > 25mm)
    HIGHLIGHT_RED = PatternFill(start_color='FF4C4C', end_color='FF4C4C', fill_type='solid')
    for row_idx in range(2, ws_missing.max_row + 1):
        liq_cap_val = liq_cap_values[row_idx - 2]
        if liq_cap_val is not None and liq_cap_val > SIGNIFICANT_MV_THRESHOLD:
            ws_missing.cell(row=row_idx, column=deal_name_display_col).fill = HIGHLIGHT_GRAY
        if ws_missing.cell(row=row_idx, column=aat_irr_display_col).value is None:
            ws_missing.cell(row=row_idx, column=aat_irr_display_col).fill = HIGHLIGHT_RED
        if ws_missing.cell(row=row_idx, column=duration_aat_display_col).value is None:
            ws_missing.cell(row=row_idx, column=duration_aat_display_col).fill = HIGHLIGHT_RED

    print(f"  - 'Missing AAT Data': {len(missing_rows)} deals listed")


def highlight_and_group_summary(ws: Worksheet, current_date: str) -> None:
    """
    Highlight significant deals and group/hide smaller ones.

    Args:
        ws: Summary worksheet to process
        current_date: Formatted current date string

    Raises:
        KeyError: If required columns not found
    """
    # Find column indices
    deal_name_col_idx = None
    mv_col_idx = None

    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Deal Name':
            deal_name_col_idx = col[0].column
        if col[0].value == f'{current_date} MV':
            mv_col_idx = col[0].column
        if deal_name_col_idx and mv_col_idx:
            break

    if deal_name_col_idx is None or mv_col_idx is None:
        raise KeyError("'Deal Name' or market value column not found")

    # Process rows
    for row in range(2, ws.max_row + 1):
        mv_value = ws.cell(row=row, column=mv_col_idx).value
        deal_name_cell = ws.cell(row=row, column=deal_name_col_idx)

        if mv_value is not None and mv_value > SIGNIFICANT_MV_THRESHOLD:
            deal_name_cell.fill = HIGHLIGHT_GRAY
        else:
            ws.row_dimensions[row].outlineLevel = 1
            ws.row_dimensions[row].hidden = True


def main(date_str: str) -> None:
    """
    Main execution function for AAT vs ECF cross-validation report generation.

    Args:
        date_str: Date string in format 'YYYYMMDD'
    """
    try:
        # Initialize dates and paths
        current_date, last_date = initialize_dates(date_str)
        paths = get_file_paths(date_str)

        # Ensure output directory exists
        os.makedirs(paths['output_folder'], exist_ok=True)
        output_path = os.path.join(paths['output_folder'], paths['output_filename'])

        # Load and process data
        print("Loading data...")
        df = load_data(paths['aat_data'], paths['status_final'])

        print("Processing data...")
        df = process_data(df, paths['aat_pm_owner'], current_date, last_date)
        df = calculate_metrics(df, current_date)
        df = reorder_columns(df, current_date, last_date)

        print("Saving to Excel...")
        save_to_excel(df, output_path)

        # Format workbook and create additional sheets
        print("Formatting workbook...")
        wb = load_workbook(output_path)
        ws = wb['Summary']

        format_worksheet(ws, current_date)

        # Identify and create sheets for significant items
        sig_changes, sig_diffs, dur_diffs = identify_significant_changes(ws, current_date)
        create_highlighted_sheets(wb, sig_changes, sig_diffs, dur_diffs, current_date, last_date)

        # Create missing AAT data sheet
        create_missing_aat_sheet(wb, current_date)

        # Add categorization and final formatting
        add_category_column(wb, current_date)
        highlight_and_group_summary(ws, current_date)
        drop_cumulative_mv_column(wb)

        # Save and close
        wb.save(output_path)
        wb.close()

        print(f"\u2705 Report generated successfully: {output_path}")

    except FileNotFoundError as e:
        print(f"\u274c Error: {e}")
        raise
    except KeyError as e:
        print(f"\u274c Column error: {e}")
        raise
    except Exception as e:
        print(f"\u274c Unexpected error: {e}")
        raise


def find_next_version(date_str: str, versioned_folder: str) -> int:
    """
    Find the next version number for the given date.

    Args:
        date_str: Date string in format 'YYYYMMDD'
        versioned_folder: Folder containing versioned files

    Returns:
        Next version number (1 if no existing versions found)
    """
    try:
        all_files = [f for f in os.listdir(versioned_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
    except Exception as e:
        print(f"  [Warning] Cannot read versioned folder {versioned_folder}: {e}")
        return 1

    # Pattern: date (8 digits) followed by .v and version number
    pattern = rf'{date_str}\.v(\d+)'

    max_version = 0
    for filename in all_files:
        match = re.search(pattern, filename)
        if match:
            version = int(match.group(1))
            max_version = max(max_version, version)

    return max_version + 1


def save_versioned_copy(output_path: str, date_str: str, versioned_folder: str) -> None:
    """
    Save a versioned copy of the output file to the versioned files folder.

    Args:
        output_path: Path to the generated output file
        date_str: Date string in format 'YYYYMMDD'
        versioned_folder: Folder to save versioned files
    """
    try:
        # Ensure versioned folder exists
        os.makedirs(versioned_folder, exist_ok=True)

        # Find next version number
        next_version = find_next_version(date_str, versioned_folder)

        # Create versioned filename
        versioned_filename = f'AAT vs ECF {date_str}.v{next_version}.xlsx'
        versioned_path = os.path.join(versioned_folder, versioned_filename)

        # Copy file
        shutil.copy2(output_path, versioned_path)
        print(f"  [OK] Versioned copy saved: {versioned_filename}")

    except Exception as e:
        print(f"  [Warning] Failed to save versioned copy: {e}")


def run_cross_validation(date_str: str) -> None:
    """
    Public interface to run the cross-validation report generation.

    This function can be called from main.py or other modules.

    Args:
        date_str: Date string in format 'YYYYMMDD' (e.g., '20251130')
    """
    main(date_str)

    # Save versioned copy
    paths = get_file_paths(date_str)
    output_path = os.path.join(paths['output_folder'], paths['output_filename'])
    save_versioned_copy(output_path, date_str, config.VERSIONED_FILES_FOLDER)


if __name__ == "__main__":
    # Allow running this module directly with a default date
    # Change the date here if running this module standalone
    DEFAULT_DATE = '20251130'
    main(DEFAULT_DATE)
