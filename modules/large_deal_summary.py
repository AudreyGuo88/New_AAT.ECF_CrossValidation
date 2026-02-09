"""
Large Deal Summary for Dave

This module creates a summary report by:
1. Copying the source file to output folder
2. Renaming to "Large Deals Summary for Dave.xlsx"
3. Keeping only specified columns in Summary tab
4. Removing rows with "CoreWeave" in Deal Name
5. Adding % LC column and calculating percentages
6. Highlighting top 10 Deal Names
"""

import os
import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from copy import copy

# Add parent directory to path to import config
parent_dir = Path(__file__).parent.parent
sys.path.insert(0, str(parent_dir))

import config

# ===== Configuration =====
SOURCE_FOLDER = config.AAT_ECF_SUMMARY_REPORT
OUTPUT_FOLDER = config.LARGE_DEAL_SUMMARY_FOLDER
OUTPUT_FILENAME = 'Large Deals Summary for Dave.xlsx'

# Columns to keep (will be formatted with date)
COLUMNS_TO_KEEP = [
    'Deal Name',
    'Sr. Portfolio Manager',
    '{date} AAT IRR',
    '{date} ECF IRR',
    'AAT&ECF IRR Diffs',
    'Duration AAT',
    'Duration ECF',
    'Duration Diffs',
    'Liq Cap',
    'Category'
]

# Highlight color for top 10
HIGHLIGHT_TOP10 = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")


def run_large_deal_summary(date_str: str) -> None:
    """
    Main function to create large deal summary report.

    Args:
        date_str: Date string in format 'YYYYMMDD'
    """
    print("\n" + "=" * 80)
    print("Module: Large Deal Summary for Dave")
    print("=" * 80)
    print(f"\nProcessing date: {date_str}")

    # Convert date to formatted date
    from datetime import datetime
    dt = datetime.strptime(date_str, '%Y%m%d')
    current_date = f"{dt.month}/{dt.day}/{str(dt.year)[-2:]}"

    # Step 1: Find and copy source file
    print("\n[1/5] Finding and copying source file...")
    source_filename = f'AAT vs ECF {date_str}.xlsx'
    source_file = os.path.join(SOURCE_FOLDER, source_filename)

    if not os.path.exists(source_file):
        print(f"  [Error] File not found: {source_file}")
        return

    # Ensure output folder exists
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Copy and rename
    dest_file = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    shutil.copy2(source_file, dest_file)
    print(f"  - Copied to: {OUTPUT_FILENAME}")

    # Step 2: Open workbook and process Summary tab
    print("\n[2/5] Processing Summary tab...")
    wb = load_workbook(dest_file)

    if 'Summary' not in wb.sheetnames:
        print("  [Error] 'Summary' tab not found")
        wb.close()
        return

    ws = wb['Summary']

    # Find the data table header row (skip pivot table at top)
    # Look for the row containing "Deal Name" header
    data_start_row = None
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col_idx).value == 'Deal Name':
                data_start_row = row_idx
                break
        if data_start_row:
            break

    if not data_start_row:
        print("  [Error] 'Deal Name' header not found in Summary tab")
        wb.close()
        return

    print(f"  - Found data table starting at row {data_start_row}")

    # Get headers from data table
    original_headers = [ws.cell(row=data_start_row, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  - Data table columns: {[h for h in original_headers if h]}")

    # Step 3: Delete unnecessary columns (only from data table)
    print("\n[3/5] Removing unnecessary columns from data table...")

    # Build column names with date
    columns_to_keep = [col.replace('{date}', current_date) for col in COLUMNS_TO_KEEP]

    # Delete columns (iterate backwards to avoid index shifting)
    for col_idx in range(ws.max_column, 0, -1):
        header = ws.cell(row=data_start_row, column=col_idx).value
        if header not in columns_to_keep:
            ws.delete_cols(col_idx)

    print(f"  - Kept columns: {[ws.cell(row=data_start_row, column=c).value for c in range(1, ws.max_column + 1)]}")

    # Step 4: Remove CoreWeave rows
    print("\n[4/5] Removing CoreWeave deals...")

    # Find Deal Name column (after column deletion)
    deal_name_col = None
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=data_start_row, column=col_idx).value
        if header_value == 'Deal Name':
            deal_name_col = col_idx
            break

    if not deal_name_col:
        # Debug: print all remaining headers
        headers = [ws.cell(row=data_start_row, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  [Debug] Remaining headers: {headers}")
        print("  [Error] 'Deal Name' column not found after removing columns")
        wb.close()
        return

    # Delete rows with CoreWeave (iterate backwards, only from data rows)
    rows_deleted = 0
    for row_idx in range(ws.max_row, data_start_row, -1):
        deal_name = ws.cell(row=row_idx, column=deal_name_col).value
        if deal_name and 'CoreWeave' in str(deal_name):
            ws.delete_rows(row_idx)
            rows_deleted += 1

    print(f"  - Removed {rows_deleted} CoreWeave deals")

    # Step 5: Add % LC column and calculate percentages
    print("\n[5/5] Adding % LC column and highlighting top 10...")

    # Find Liq Cap column
    liq_cap_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=data_start_row, column=col_idx).value == 'Liq Cap':
            liq_cap_col = col_idx
            break

    if not liq_cap_col:
        print("  [Error] 'Liq Cap' column not found")
        wb.close()
        return

    # Insert % LC column after Liq Cap
    pct_lc_col = liq_cap_col + 1
    ws.insert_cols(pct_lc_col)

    # Set header value and copy formatting from Liq Cap header
    pct_lc_header_cell = ws.cell(row=data_start_row, column=pct_lc_col)
    pct_lc_header_cell.value = '% LC'

    # Copy header formatting from Liq Cap column
    liq_cap_header_cell = ws.cell(row=data_start_row, column=liq_cap_col)
    if liq_cap_header_cell.has_style:
        pct_lc_header_cell.font = copy(liq_cap_header_cell.font)
        pct_lc_header_cell.border = copy(liq_cap_header_cell.border)
        pct_lc_header_cell.fill = copy(liq_cap_header_cell.fill)
        pct_lc_header_cell.number_format = copy(liq_cap_header_cell.number_format)
        pct_lc_header_cell.protection = copy(liq_cap_header_cell.protection)
        pct_lc_header_cell.alignment = copy(liq_cap_header_cell.alignment)

    # Calculate total Liq Cap (only from data rows)
    total_liq_cap = 0
    for row_idx in range(data_start_row + 1, ws.max_row + 1):
        liq_cap_value = ws.cell(row=row_idx, column=liq_cap_col).value
        if liq_cap_value and isinstance(liq_cap_value, (int, float)):
            total_liq_cap += liq_cap_value

    # Calculate and write % LC (only to data rows)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row_idx in range(data_start_row + 1, ws.max_row + 1):
        liq_cap_value = ws.cell(row=row_idx, column=liq_cap_col).value
        cell = ws.cell(row=row_idx, column=pct_lc_col)
        if liq_cap_value and isinstance(liq_cap_value, (int, float)) and total_liq_cap > 0:
            pct = (liq_cap_value / total_liq_cap) * 100
            cell.value = pct / 100  # Store as decimal
            cell.number_format = '0.00%'
        # Set center alignment for all data cells
        cell.alignment = center_alignment

    # Highlight top 10 Deal Names (first 10 data rows)
    data_row_count = ws.max_row - data_start_row
    top_10_count = min(10, data_row_count)
    for row_idx in range(data_start_row + 1, data_start_row + 1 + top_10_count):
        ws.cell(row=row_idx, column=deal_name_col).fill = HIGHLIGHT_TOP10

    print(f"  - Added % LC column")
    print(f"  - Highlighted top {top_10_count} Deal Names")

    # Save and close
    wb.save(dest_file)
    wb.close()

    print(f"\n{'=' * 80}")
    print(f"[SUCCESS] Large Deal Summary created!")
    print(f"  - File: {OUTPUT_FILENAME}")
    print(f"  - Location: {OUTPUT_FOLDER}")
    print(f"  - Total deals: {ws.max_row - data_start_row}")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    # For standalone testing
    DEFAULT_DATE = '20251130'
    run_large_deal_summary(DEFAULT_DATE)
